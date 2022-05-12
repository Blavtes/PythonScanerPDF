import fitz  # pip install PyMuPDF
import os
from IPython.display import Image
import pandas as pd
from os import path 
from openpyxl import load_workbook 
from pathlib import Path
import openpyxl

import xlwt
import xlrd

# 批量获取pdf 指定区域内容 写入exls
# mac os 使用python3 运行
#    python3 -m pip install install pandas==1.2.5
#    python3 -m pip install PyMuPDF
#    python3 -m pip install Image
#    python3 -m pip install IPython
#    python3 -m pip install pandas
#    python3 -m pip install openpyxl
#    python3 -m pip install ExcelWriter
#    python3 -m pip install xlwt
#    python3 -m pip install xlrd

class Excel_Operate():

    # 批量获取的结果，一次写入，追加写入
    result = []
    """
    openpyxl 只支持xlsx
    优点能够读写大文件
    """
    wb = openpyxl.Workbook()

    @classmethod 
    #老模板数据
    def getOldPDF(cls,url):

        pdf_path = url
        if not os.path.exists("imgs"):
            os.mkdir("imgs")

        
        with fitz.open(pdf_path) as pdfDoc:
            for i in range(1):
                page_num = i+1
                print("--------------", page_num, "--------------")
                file = Path(pdf_path).stem
                fileName = file +'.pdf'
                page = pdfDoc[i]
                mat = fitz.Matrix(1.3, 1.3)  
                rect = page.rect

                a_text = Excel_Operate.getImageData(page,file,0.25*rect.width, 0.18*rect.height,
                                 rect.width*0.5, 170)
                b_text = Excel_Operate.getImageData(page,file,0.22*rect.width, 0.2*rect.height,
                                 rect.width*0.5, 200)
                c_text = Excel_Operate.getImageData(page,file,0.69*rect.width, 0.1*rect.height,
                                 rect.width*1, 110)
                d_text = Excel_Operate.getImageData(page,file,0.7*rect.width, 0.125*rect.height,
                                 rect.width*1, 130)

                cls.result.append((a_text, b_text,c_text,d_text,fileName))

    @classmethod
     #截取图片数据
    def getImageData(cls,page,fileName,x,y,width,height):
        mat = fitz.Matrix(1.5, 1.5)  
        clip = fitz.Rect(x, y, width, height)  
        pix = page.getPixmap(matrix=mat.preRotate(0),
                             alpha=False, clip=clip)  
        #pix.writeImage(f"imgs/{fileName}_a.png")
        img1 = pix.getImageData()
        #display(Image(img1))
        a_text = page.getText(clip=clip)
        print(a_text)
        return a_text

    @classmethod
     #有电子签章
    def getNewPDF(cls,url,tag):

        pdf_path = url
        if not os.path.exists("imgs"):
            os.mkdir("imgs")

        with fitz.open(pdf_path) as pdfDoc:
            for i in range(1):
                page_num = i+1
                print("---------new-----", page_num, "--------------")
                file = Path(pdf_path).stem
                fileName = file +'.pdf'
                page = pdfDoc[i]
                mat = fitz.Matrix(1.5, 1.5)  
                rect = page.rect

                #获取姓名
               
                a_text = Excel_Operate.getImageData(page,file,0.18*rect.width, 0.09*rect.height,
                                 rect.width*0.5, 95)
                #获取身份证
                b_text = Excel_Operate.getImageData(page,file,0.15*rect.width, 0.1*rect.height,
                                 rect.width*0.5, 110)
                #获取协议编号
                c_text = Excel_Operate.getImageData(page,file,0.6*rect.width, 0.04*rect.height,
                                 rect.width*1, 55)
                #获取编号
                d_text = Excel_Operate.getImageData(page,file,0.6*rect.width, 0.06*rect.height,
                                 rect.width*1, 80)
                if (tag == 2) :
                     d_text = Excel_Operate.getImageData(page,file,0.75*rect.width, 0.06*rect.height,
                                 rect.width*1, 80)
                """
                pix = page.getPixmap(matrix=mat.preRotate(0),
                 #                    alpha=False, clip=clip)  
                # 储存截取得图片
                 pix.writeImage(f"imgs/{fileName}_d.png")
                img4 = pix.getImageData()
                # display(Image(img1))
                d_text = page.getText(clip=clip)
                tmp = pd.DataFrame(d_text.splitlines(), columns=["a"])
                tmp["b"] = (tmp.a.str[:2]).astype("category")
                tmp.b.cat.set_categories(
                    ['An', 're', 'vi', 'Ma', 'co', 'VC', 'ES'], inplace=True)
                tmp.sort_values('b', inplace=True)
                d_text = '\n'.join(tmp.a.to_list())
                print(d_text)
                """

                if '广金服编号' in d_text:
                     Excel_Operate.getNewPDF(url,2)
                else :
                    cls.result.append([a_text, b_text,c_text,d_text,fileName])
    """
    @classmethod # 监测elxs 是否存在
    def createElxs(cls,wb_name):
        try:
            with xlrd.open_workbook(wb_name) as wb:
                sh1 = wb.sheet_by_index(0)
                for sh in wb.sheets():
                    for r in range(sh.nrows):
                        # 输出指定行
                        print(sh.row(r))
        except FileNotFoundError:
            # 创建工作簿
            wb = xlwt.Workbook(wb_name)
    """
    @classmethod 
    #搜索文件函数
    def scaner_file(cls,url , elxsName):
            #Excel_Operate.createElxs(elxsName)
            #遍历当前路径下所有文件
            file  = os.listdir(url)
            for f in file:
                #字符串拼接
                real_url = path.join (url , f)
                #打印出来
                print(real_url)
                file = Path(real_url).stem
                print(file)
                if '.pdf' in real_url:
                    ll = file[3:11]
                    if (int(ll) < 20180127):
                        Excel_Operate.getOldPDF(real_url)
                    else :
                        Excel_Operate.getNewPDF(real_url,1)
                # break
            print(cls.result)
            """
            # 写入数据到xlsx
             data = pd.DataFrame(result, columns=["姓名","身份证号","协议编号","广金所编号","文件名"])
             writer = pd.ExcelWriter(elxsName,mode='a', engine='openpyxl',if_sheet_exists='new')
             data.to_excel(writer, sheet_name='sheet1')
             writer.save()
             writer.close()
            """
            #写入数据
            Excel_Operate.add_to_data(elxsName, data=cls.result)

    #调用自定义函数
    #scaner_file("./test" , 'result.xlsx') 

    @classmethod
    def create(cls, ex_path, name, sheet_name="Sheet1",rowName=[], sheet_names=[]):
        """
        :param ex_path:保存路径
        :param name: excel名称
        :param sheet_name: 默认sheet名称
        :param sheet_names:其它sheet名称
        :return:
        """
        if not os.path.exists(ex_path):
            os.makedirs(ex_path)

        ex_path_name = os.path.join(ex_path, name)
        if os.path.exists(ex_path_name):
            return True

        ws1 = cls.wb.active  # 默认表sheet1
        if sheet_name:
            ws1.title = sheet_name

        for name in sheet_names:  # 定义其它sheet页
            cls.wb.create_sheet(str(name))

        cls.wb.save(ex_path_name)
        Excel_Operate.add_to_data(name, data=rowName,sheetname=sheet_name)

    @classmethod
    def add_to_data(cls, ex_path_name, data, sheetname="Sheet1"):
        """

        :param ex_path_name:
        :param data: [[1, 2, 3], [4, 5, 6]]
        :return:
        """

        if not os.path.exists(ex_path_name):
            return {"status": False, "message": "文件不存在"}

        wb = openpyxl.load_workbook(ex_path_name)
        # sheetnames = wb.sheetnames  # [u'Sheet'] 获取所有sheet页list
        print("------- sheetname  -----" + sheetname)
        ws = wb[sheetname]  # 选取第一个sheet页

        for x in data:
            ws.append(x)
        wb.save(ex_path_name)

    @classmethod
    def read_data(cls, ex_path_name, is_col, is_row=True):
        """

        :param ex_path_name:
        :param is_col: 以列形式返回
        :param is_row: 以行形式返回
        :return:
        """

        wb = openpyxl.load_workbook(ex_path_name)
        # 获取全部表名
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        # 表总行数max_row
        max_row = ws.max_row
        # 表总列数
        max_col = ws.max_column

        row_data = []  # 行数据
        col_data = []  # 列数据

        if is_row:
            for i in range(1, max_row + 1):
                t_data = []
                for x in range(1, max_col + 1):
                    # 获取表中x行1列的值
                    cell_data = ws.cell(column=x, row=i).value
                    t_data.append(cell_data)
                row_data.append(t_data)
            return row_data

        if is_col:
            for i in range(1, max_col + 1):
                t_data = []
                for x in range(1, max_row + 1):
                    # 获取表中x行1列的值
                    cell_data = ws.cell(row=x, column=i).value
                    t_data.append(cell_data)
                col_data.append(t_data)
            return col_data
        return []


if __name__ == "__main__":
    elxsName = "result.xlsx"
    Excel_Operate.create("./", elxsName,rowName=[["姓名","身份证号","协议编号","编号","文件名"]])
    # Excel_Operate.add_to_data("./result.xlsx", data=[[1,2,3,4,5]], sheetname=u"协议")
    Excel_Operate.scaner_file("./all" ,elxsName)


